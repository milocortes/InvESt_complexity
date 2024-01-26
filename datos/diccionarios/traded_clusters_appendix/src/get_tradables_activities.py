# Cargamos paquetes
import pandas as pd 
from openpyxl import load_workbook
import numpy as np
import os 

# Definimos rutas
FILE_PATH = os.getcwd()
DATA_PATH = os.path.abspath(os.path.join(FILE_PATH, "..", "datos"))
OUTPUT_PATH = os.path.abspath(os.path.join(FILE_PATH, "..", "output"))
DATA_FILE_PATH = os.path.join(DATA_PATH, "clases_transables","Traded Clusters Appendix.xlsx")

# Obtenemos los nombres de las pestañas del archivo
wb = load_workbook(DATA_FILE_PATH, read_only=True, keep_links=False)
data_sheetnames = wb.sheetnames[3:55]
wb.close()

# Cargamos archivo
traded_cluster = {}

for cluster in data_sheetnames:
    print(cluster)
    cluster_data = pd.read_excel(DATA_FILE_PATH, sheet_name=cluster, skiprows =12, usecols = "B,C,F")
    cluster_data.columns = ["naics", "naics_name", "subcluster_name"]
    traded_cluster[cluster] = cluster_data

del traded_cluster['Househld Textiles & Leather Pro']

# Cargamos las clases de los naics 2007, 2012 y 2017 para verificar en qué clasificación se encuentran los datos
# y, si hubo algun cambio, hacer el crosswalk a naics 2017

naics_07 = pd.read_excel(os.path.join(DATA_PATH, "naics-6-code-02-07-12-17", "naics07_6.xls"), skiprows=1)
naics_07.columns = ["naics07", "naics_name07"]

naics_12 = pd.read_excel(os.path.join(DATA_PATH, "naics-6-code-02-07-12-17", "6-digit_2012_Codes.xls"), skiprows=1)
naics_12.columns = ["naics12", "naics_name12"]

naics_17 = pd.read_excel(os.path.join(DATA_PATH, "naics-6-code-02-07-12-17","6-digit_2017_Codes.xlsx"), skiprows=1).iloc[:,[0,1]]
naics_17.columns = ["naics17", "naics_name17"]


# Obtenemos todas las clases que son tradables
clases_transables = []
for k,v in traded_cluster.items():
    clases_transables += [i for i in v["naics"].to_list() if i!="NAICS"]

clases_transables = [i for i in clases_transables if not np.isnan(i)]
clases_transables = list(set(clases_transables))
clases_transables.sort()

# Las clases pertenecen a la clasificación 2007
print(len(clases_transables))
print(len(set(clases_transables).intersection(naics_07.naics07)))

### Cargamos crosswalks entre naics 2007 --> naics 2012 y naics 2012 ---> naics 2017
## CORRESPONDENCIA naics 2007 --> naics 2012

naics_07_12_tuple = pd.read_excel(os.path.join(DATA_PATH, "concordances_cw", "2007_to_2012_NAICS.xls"), sheet_name = "2007 to 2012 NAICS U.S.", skiprows = 2)\
                .iloc[:,[0,1,2,3]][["2007 NAICS Code", "2012 NAICS Code"]]\
                .to_records(index = False)

naics_07_12 = {i:[] for i,j in naics_07_12_tuple}

for i,j in naics_07_12_tuple:
    naics_07_12[i].append(j)

naics_07_12_one_to_many = {i:j for i,j in naics_07_12.items() if len(j)>1}
naics_07_12_one_to_one = {i:j[0] for i,j in naics_07_12.items() if len(j)==1}

## Reemplazamos las relaciones uno a uno
for cluster, datos_cluster in traded_cluster.items():
    datos_cluster_nuevo = datos_cluster.copy()
    datos_cluster_nuevo["naics"] = datos_cluster_nuevo["naics"].replace(naics_07_12_one_to_one)
    traded_cluster[cluster] = datos_cluster_nuevo

## Reemplazamos las relaciones uno a muchos
for cluster, datos_cluster in traded_cluster.items():
    for i,j in naics_07_12_one_to_many.items():
        if i in datos_cluster.naics.to_list():  
            print(f"Si sale actividad {i} en cluster {cluster}")

            parcial_apertura = pd.DataFrame({"naics": j}).merge(right=datos_cluster.query(f"naics=={i}")[["naics_name", "subcluster_name"]], how = "cross")
            datos_cluster_nuevo = datos_cluster.query(f"naics!={i}")
            datos_cluster_nuevo = pd.concat([datos_cluster_nuevo, parcial_apertura], ignore_index = True)
            traded_cluster[cluster] = datos_cluster_nuevo

## Verificamos que todas las clases se encuentren en naics 2012
clases_transables = []
for k,v in traded_cluster.items():
    clases_transables += [i for i in v["naics"].to_list() if i!="NAICS"]

clases_transables = [i for i in clases_transables if not np.isnan(i)]
clases_transables = list(set(clases_transables))
clases_transables.sort()

print(len(clases_transables))
print(len(set(clases_transables).intersection(naics_12.naics12)))

## CORRESPONDENCIA naics 2012 --> naics 2017
naics_12_17_tuple = pd.read_excel(os.path.join(DATA_PATH, "concordances_cw", "2012_to_2017_NAICS.xlsx"), sheet_name = "2012 to 2017 NAICS U.S.", skiprows = 2)\
                .iloc[:,[0,1,2,3]][["2012 NAICS Code", "2017 NAICS Code"]]\
                .to_records(index = False)

naics_12_17 = {i:[] for i,j in naics_12_17_tuple}

for i,j in naics_12_17_tuple:
    naics_12_17[i].append(j)

naics_12_17_one_to_many = {i:j for i,j in naics_12_17.items() if len(j)>1}
naics_12_17_one_to_one = {i:j[0] for i,j in naics_12_17.items() if len(j)==1}


## Reemplazamos las relaciones uno a uno
for cluster, datos_cluster in traded_cluster.items():
    datos_cluster_nuevo = datos_cluster.copy()
    datos_cluster_nuevo["naics"] = datos_cluster_nuevo["naics"].replace(naics_12_17_one_to_one)
    traded_cluster[cluster] = datos_cluster_nuevo

## Reemplazamos las relaciones uno a muchos
for cluster, datos_cluster in traded_cluster.items():
    for i,j in naics_12_17_one_to_many.items():
        if i in datos_cluster.naics.to_list():  
            print(f"Si sale actividad {i} en cluster {cluster}")

            parcial_apertura = pd.DataFrame({"naics": j}).merge(right=datos_cluster.query(f"naics=={i}")[["naics_name", "subcluster_name"]], how = "cross")
            datos_cluster_nuevo = datos_cluster.query(f"naics!={i}")
            datos_cluster_nuevo = pd.concat([datos_cluster_nuevo, parcial_apertura], ignore_index = True)
            traded_cluster[cluster] = datos_cluster_nuevo

## Verificamos que todas las clases se encuentren en naics 2012
clases_transables = []
for k,v in traded_cluster.items():
    clases_transables += [i for i in v["naics"].to_list() if i!="NAICS"]

clases_transables = [i for i in clases_transables if not np.isnan(i)]
clases_transables = list(set(clases_transables))
clases_transables.sort()

print(len(clases_transables))
print(len(set(clases_transables).intersection(naics_17.naics17)))

## Guardamos las clases transables en un diccionario para guardarlo como json
clases_transables.remove(541711)
clases_transables = {"clases_transables" : clases_transables}

import json 

json.dump(clases_transables, open(os.path.join(OUTPUT_PATH, "clases_transables_naics_2017.json"),"w"))